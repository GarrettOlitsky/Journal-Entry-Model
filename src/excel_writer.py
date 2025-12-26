from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def write_je_xlsx(payload: dict, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entry"

    ws.append(["Date", "Account", "Debit", "Credit", "Memo"])

    for line in payload["lines"]:
        ws.append([payload["date"], line["account"], line["debit"], line["credit"], payload["memo"]])

    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 42
    ws.column_dimensions[get_column_letter(3)].width = 12
    ws.column_dimensions[get_column_letter(4)].width = 12
    ws.column_dimensions[get_column_letter(5)].width = 36

    wb.save(out_path)
